"""Import utilities for batch URL/file import and deduplication."""

from __future__ import annotations

from pathlib import Path
from urllib.parse import urlparse

from flowscribe.app.models import SourceSpec
from flowscribe.queue.models import QueueItem


def parse_urls_from_text(text: str) -> list[str]:
    urls: list[str] = []
    for line in text.splitlines():
        candidate = line.strip()
        if not candidate or candidate.startswith("#"):
            continue
        parsed = urlparse(candidate)
        if parsed.scheme in ("http", "https") and parsed.netloc:
            urls.append(candidate)
    return urls


def import_urls_from_txt(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8", errors="replace")
    return parse_urls_from_text(text)


def import_urls_from_csv(path: Path) -> list[str]:
    import csv

    urls: list[str] = []
    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            candidate = row[0].strip()
            parsed = urlparse(candidate)
            if parsed.scheme in ("http", "https") and parsed.netloc:
                urls.append(candidate)
    return urls


def import_urls_from_xlsx(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    urls: list[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        cell_value = row[0]
        if cell_value is None:
            continue
        candidate = str(cell_value).strip()
        parsed = urlparse(candidate)
        if parsed.scheme in ("http", "https") and parsed.netloc:
            urls.append(candidate)
    wb.close()
    return urls


def import_urls_from_file(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return import_urls_from_txt(path)
    if suffix == ".csv":
        return import_urls_from_csv(path)
    if suffix == ".xlsx":
        return import_urls_from_xlsx(path)
    raise ValueError(f"Unsupported import file type: {suffix}")


def deduplicate_sources(
    new_sources: list[SourceSpec],
    existing_items: list[QueueItem],
) -> list[SourceSpec]:
    existing_keys: set[str] = set()
    for item in existing_items:
        if item.status in ("pending", "running"):
            existing_keys.add(f"{item.source.kind}:{item.source.value}")
    return [
        source
        for source in new_sources
        if f"{source.kind}:{source.value}" not in existing_keys
    ]
